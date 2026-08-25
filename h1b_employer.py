#!/usr/bin/env python3
"""DOL LCA employer index + H-1B / green-card validation helpers for quickjobs."""

from __future__ import annotations

import json
import random
import re
import urllib.request
from pathlib import Path
from typing import Any, Callable

DOL_LCA_URL = (
    "https://www.dol.gov/sites/dolgov/files/ETA/oflc/pdfs/"
    "LCA_Disclosure_Data_FY{fy}_Q{quarter}.xlsx"
)

H1B_POSITIVE_SPONSOR_PHRASES = (
    "visa sponsorship is available",
    "visa sponsorship available",
    "visa sponsorship or transfer is available",
    "offer visa transfer",
    "able to offer visa transfer",
    "open to visa transfer",
    "visa transfer available",
    "willing to facilitate visa transfers",
    "sponsor employment visas",
    "proud to sponsor",
    "will sponsor h-1b",
    "will sponsor h1b",
    "h-1b sponsorship available",
    "h1b sponsorship available",
    "h-1b visa sponsorship",
    "h1b visa sponsorship",
    "sponsors work visas",
    "sponsor work visas",
    "provide visa sponsorship",
    "provide immigration sponsorship",
    "immigration sponsorship available",
    "eligible for visa sponsorship",
    "we sponsor visas",
    "we will sponsor",
    "sponsorship is available",
)

H1B_NEGATIVE_NO_SPONSOR_PHRASES = (
    "unable to sponsor or take over sponsorship",
    "not eligible for immigration sponsorship",
    "does not offer visa sponsorship",
    "role does not offer visa sponsorship",
    "this role does not offer visa sponsorship",
    "not offer visa sponsorship",
    "no visa sponsorship",
    "will not sponsor",
    "cannot sponsor",
    "can not sponsor",
    "not able to sponsor",
    "unable to sponsor",
    "does not sponsor",
    "do not sponsor",
    "without sponsorship",
    "visa sponsorship is not available",
    "sponsorship is not available",
    "not eligible for visa sponsorship",
    "not eligible for sponsorship",
    "cannot provide visa sponsorship",
    "will not provide visa sponsorship",
    "does not provide visa sponsorship",
    "no sponsorship available",
    "us citizens only",
    "u.s. citizens only",
    "must be a us citizen",
    "must be authorized to work in the united states without sponsorship",
    "must be authorized to work in the us without sponsorship",
    "must be authorized to work for any employer without sponsorship",
    "work authorization without sponsorship",
)

_VISA_SPONSOR_DENIAL_CONTEXT = (
    "without sponsorship",
    "no sponsorship",
    "not sponsor",
    "cannot sponsor",
    "will not sponsor",
    "unable to sponsor",
    "not eligible for",
)

_EMPLOYER_INDEX: dict[str, dict[str, Any]] | None = None
_WAGE_INDEX: dict[str, dict[str, Any]] | None = None
_COMPANY_LOOKUP_CACHE: dict[str, dict[str, Any]] = {}

# Longest-first needles for LCA job-title wage buckets (tech/platform roles).
LCA_TITLE_NEEDLES: tuple[str, ...] = (
    "site reliability engineer",
    "software development engineer",
    "senior software engineer",
    "staff software engineer",
    "principal software engineer",
    "machine learning engineer",
    "software engineer",
    "devops engineer",
    "platform engineer",
    "reliability engineer",
    "security engineer",
    "systems engineer",
    "infrastructure engineer",
    "cloud engineer",
    "data engineer",
    "network engineer",
    "production engineer",
    "support engineer",
    "solutions engineer",
)

_WAGE_RESERVOIR_N = 400
_TITLE_RESERVOIR_N = 200
_MIN_TITLE_SAMPLES = 3
_MIN_EMPLOYER_SAMPLES = 5


def profile_wants_h1b_validation(cfg: dict[str, Any]) -> bool:
    status = str((cfg.get("profile") or {}).get("resident_status") or "citizen").strip().lower()
    return status == "h1b"


def profile_wants_green_card_validation(cfg: dict[str, Any]) -> bool:
    status = str((cfg.get("profile") or {}).get("resident_status") or "citizen").strip().lower()
    return status == "green_card"


# Board text-filter presets (one Doesn't contain chip per phrase).
# Substring match on title + description; keep visa-specific (avoid bare
# "visa", "sponsorship", or "work authorization" — too many false positives).
H1B_BOARD_FILTER_EXCLUDE_TERMS = H1B_NEGATIVE_NO_SPONSOR_PHRASES

GREEN_CARD_BOARD_FILTER_EXCLUDE_TERMS = (
    "citizenship is strictly required",
    "us citizenship is strictly required",
    "us citizenship required",
    "must be a us citizen",
    "us citizens only",
    "applicants for this position must be us citizens",
)


def profile_resident_status(cfg: dict[str, Any]) -> str:
    return str((cfg.get("profile") or {}).get("resident_status") or "citizen").strip().lower()


def profile_default_text_filters(cfg: dict[str, Any]) -> list[dict[str, str]]:
    """Default bottom-bar text filter chips for visa / green-card profiles."""
    status = profile_resident_status(cfg)
    if status == "h1b":
        return [
            {"mode": "not", "text": term}
            for term in H1B_BOARD_FILTER_EXCLUDE_TERMS
        ]
    if status == "green_card":
        return [
            {"mode": "not", "text": term}
            for term in GREEN_CARD_BOARD_FILTER_EXCLUDE_TERMS
        ]
    return []


def profile_default_filter_scope(cfg: dict[str, Any]) -> dict[str, bool]:
    """Checkbox defaults for profile text-filter presets."""
    status = profile_resident_status(cfg)
    if status == "h1b":
        return {"title": True, "description": True}
    if status == "green_card":
        return {"title": True, "description": True}
    return {"title": True, "description": False}


def _load_no_visa_sponsor_company_ids(root: Path | None = None) -> list[str]:
    """Employers that do not sponsor work visas (config snippet or portable helper)."""
    try:
        from no_visa_sponsor_company_ids import load_no_visa_sponsor_company_ids

        return load_no_visa_sponsor_company_ids(root or Path(__file__).resolve().parent)
    except ImportError:
        pass
    base = root or Path(__file__).resolve().parent
    snippet = base / "config" / "no-visa-sponsor-company-ids.json"
    if snippet.is_file():
        payload = json.loads(snippet.read_text(encoding="utf-8"))
        raw = payload.get("company_ids") if isinstance(payload, dict) else payload
        return sorted(str(cid) for cid in (raw or []) if str(cid).strip())
    return []


def employer_index_ready(cache_root: Path) -> bool:
    """True when employer-index.json exists and loaded with at least one employer."""
    path = employer_index_path(cache_root)
    if not path.is_file():
        return False
    return bool(_load_employer_index(cache_root))


def company_ids_exclude_no_dol_visa_filers(
    companies: list[dict[str, Any]],
    cache_root: Path,
) -> list[str]:
    """Company IDs with no DOL LCA filings (requires a built employer index)."""
    if not employer_index_ready(cache_root):
        return []
    excludes: list[str] = []
    for co in companies:
        cid = str(co.get("id") or "").strip()
        if not cid:
            continue
        name = str(co.get("name") or co.get("title") or "").strip()
        meta = lookup_company_h1b_meta(cid, name, cache_root=cache_root)
        if not meta.get("filer"):
            excludes.append(cid)
    return sorted(set(excludes))


def profile_default_company_ids_exclude(
    cfg: dict[str, Any],
    companies: list[dict[str, Any]] | None = None,
    cache_root: Path | None = None,
    *,
    warn: Callable[[str], None] | None = None,
) -> list[str]:
    """Default company_ids_exclude additions for visa-holder profiles."""
    if profile_resident_status(cfg) != "h1b":
        return []
    out = set(_load_no_visa_sponsor_company_ids())
    if companies is None or cache_root is None:
        return sorted(out)
    if employer_index_ready(cache_root):
        out.update(company_ids_exclude_no_dol_visa_filers(companies, cache_root))
    elif warn is not None:
        warn(
            "DOL employer index not built; skipping non-filer company excludes "
            f"(run fetch_h1b_employer_index.py — expected {employer_index_path(cache_root)})"
        )
    return sorted(out)


def _normalize_posting_text(text: str) -> str:
    normalized = str(text or "").lower()
    normalized = re.sub(r"u\.s\.", "us", normalized)
    normalized = re.sub(r"u\.s\b", "us", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


US_CITIZENSHIP_REQUIRED_PHRASES = (
    "citizenship is strictly required",
    "us citizenship is strictly required",
    "us citizenship strictly required",
    "us citizenship is required",
    "us citizenship required",
    "must be us citizens",
    "must be a us citizen",
    "must be united states citizens",
    "must be a united states citizen",
    "us citizens only",
    "applicants for this position must be us citizens",
)

_CITIZENSHIP_REQUIRED_RES = (
    re.compile(r"\bcitizenship\s+is\s+strictly\s+required\b"),
    re.compile(r"\bus\s+citizenship\s+(?:is\s+)?(?:strictly\s+)?required\b"),
    re.compile(r"\bmust\s+be\s+(?:a\s+)?us\s+citizens?\b"),
    re.compile(r"\bapplicants\b.{0,100}?\bmust\s+be\s+us\s+citizens\b", re.DOTALL),
    re.compile(r"\bus\s+citizens\s+only\b"),
)

_CITIZENSHIP_IS_REQUIRED = re.compile(r"\bcitizenship\s+is\s+required\b")
_US_CONTEXT = re.compile(r"\b(?:us|united states)\b")


def posting_text_indicates_us_citizenship_required(text: str) -> bool:
    """True when JD/title requires U.S. citizenship (green-card holders excluded)."""
    normalized = _normalize_posting_text(text)
    if not normalized:
        return False
    if any(phrase in normalized for phrase in US_CITIZENSHIP_REQUIRED_PHRASES):
        return True
    for pattern in _CITIZENSHIP_REQUIRED_RES:
        if pattern.search(normalized):
            return True
    for match in _CITIZENSHIP_IS_REQUIRED.finditer(normalized):
        start, end = match.span()
        window = normalized[max(0, start - 80) : min(len(normalized), end + 80)]
        if _US_CONTEXT.search(window):
            return True
    return False


def green_card_job_skip_reason(title: str, description_text: str) -> str | None:
    """Return skip reason when posting requires U.S. citizenship."""
    blob = f"{title}\n{description_text}"
    if posting_text_indicates_us_citizenship_required(blob):
        return "US citizenship required"
    return None


def normalize_employer_name(name: str) -> str:
    text = str(name or "").strip().lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[,.&/]+", " ", text)
    for token in (
        "incorporated",
        "corporation",
        "company",
        "limited",
        "holding",
        "holdings",
        "inc",
        "llc",
        "ltd",
        "corp",
        "co",
        "plc",
        "the",
    ):
        text = re.sub(rf"\b{token}\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def h1b_index_dir(cache_root: Path) -> Path:
    path = cache_root
    path.mkdir(parents=True, exist_ok=True)
    return path


def employer_index_path(cache_root: Path) -> Path:
    return h1b_index_dir(cache_root) / "employer-index.json"


def wage_index_path(cache_root: Path) -> Path:
    return h1b_index_dir(cache_root) / "lca-wage-index.json"


def company_h1b_cache_path(cache_root: Path, company_id: str) -> Path:
    return h1b_index_dir(cache_root) / f"{company_id}.json"


def annualize_lca_wage(amount: float, unit: str) -> float | None:
    """Convert an LCA wage amount to an approximate annual USD figure."""
    try:
        value = float(amount)
    except (TypeError, ValueError):
        return None
    if value <= 0:
        return None
    key = str(unit or "").strip().upper().replace(" ", "")
    if key in {"YEAR", "YR", "ANNUAL", "ANNUM", "Y"}:
        return value
    if key in {"HOUR", "HR", "H"}:
        return value * 2080.0
    if key in {"WEEK", "WK", "W"}:
        return value * 52.0
    if key in {"BI-WEEKLY", "BIWEEKLY", "BIWEEK", "BW"}:
        return value * 26.0
    if key in {"MONTH", "MO", "MTH", "M"}:
        return value * 12.0
    return None


def _parse_lca_amount(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw) if float(raw) > 0 else None
    text = str(raw).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    return value if value > 0 else None


def lca_row_annual_wage(
    wage_from: Any,
    wage_to: Any,
    wage_unit: Any,
) -> float | None:
    """Midpoint annual wage from LCA FROM/TO + unit columns."""
    low = _parse_lca_amount(wage_from)
    high = _parse_lca_amount(wage_to)
    if low is None and high is None:
        return None
    if low is None:
        mid = high
    elif high is None:
        mid = low
    else:
        mid = (low + high) / 2.0
    assert mid is not None
    return annualize_lca_wage(mid, str(wage_unit or "Year"))


def _reservoir_add(
    values: list[float],
    counts: dict[str, int],
    key: str,
    value: float,
    max_n: int,
) -> None:
    n = int(counts.get(key) or 0) + 1
    counts[key] = n
    if len(values) < max_n:
        values.append(value)
        return
    # Reservoir sampling so large employers stay bounded in memory.
    slot = random.randint(1, n)
    if slot <= max_n:
        values[slot - 1] = value


def _percentile(sorted_vals: list[float], pct: float) -> float:
    if not sorted_vals:
        return 0.0
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    rank = (len(sorted_vals) - 1) * pct
    lo = int(rank)
    hi = min(lo + 1, len(sorted_vals) - 1)
    frac = rank - lo
    return sorted_vals[lo] * (1.0 - frac) + sorted_vals[hi] * frac


def _wage_summary(values: list[float], *, sample_count: int | None = None) -> dict[str, Any] | None:
    clean = [float(v) for v in values if isinstance(v, (int, float)) and float(v) > 0]
    if not clean:
        return None
    clean.sort()
    n = int(sample_count if sample_count is not None else len(clean))
    p25 = int(round(_percentile(clean, 0.25)))
    p75 = int(round(_percentile(clean, 0.75)))
    if p25 > p75:
        p25, p75 = p75, p25
    return {
        "n": n,
        "p25": p25,
        "p50": int(round(_percentile(clean, 0.50))),
        "p75": p75,
        "min": int(round(clean[0])),
        "max": int(round(clean[-1])),
    }


def format_lca_salary_label(p25: int, p75: int) -> str:
    """Badge label with DOL LCA provenance (compact badge keeps the $ range)."""

    def _k(amount: int) -> str:
        if amount >= 1000:
            return f"${int(round(amount / 1000.0))}K"
        return f"${amount:,}"

    return f"{_k(p25)}\u2013{_k(p75)} · DOL LCA"


def _title_needles_for(job_title: str) -> list[str]:
    lower = re.sub(r"[^a-z0-9+\s]", " ", str(job_title or "").lower())
    lower = re.sub(r"\s+", " ", lower).strip()
    if not lower:
        return []
    return [needle for needle in LCA_TITLE_NEEDLES if needle in lower]


def compute_h1b_grade(certified: int, denied: int) -> tuple[str, str]:
    total = certified + denied
    if certified <= 0:
        return "F", "low"
    denial_rate = denied / total if total else 0.0
    if certified >= 50 and denial_rate < 0.10:
        return "A", "high"
    if certified >= 20 and denial_rate < 0.15:
        return "B", "high"
    if certified >= 10 and denial_rate < 0.20:
        return "B", "medium"
    if certified >= 3:
        return "C", "medium" if certified >= 5 else "low"
    return "D", "low"


def posting_text_indicates_no_visa_sponsorship(text: str) -> bool:
    """True when JD explicitly denies visa sponsorship or transfer."""
    lower = str(text or "").lower()
    if not lower:
        return False
    if any(phrase in lower for phrase in H1B_NEGATIVE_NO_SPONSOR_PHRASES):
        return True
    if "must be authorized to work for any employer" in lower:
        return any(marker in lower for marker in _VISA_SPONSOR_DENIAL_CONTEXT)
    return False


def posting_text_indicates_h1b_sponsor_offer(text: str) -> bool:
    lower = str(text or "").lower()
    if not lower:
        return False
    if posting_text_indicates_no_visa_sponsorship(lower):
        return False
    if any(phrase in lower for phrase in H1B_POSITIVE_SPONSOR_PHRASES):
        return True
    if "visa transfer" in lower or "offer visa transfer" in lower:
        return True
    return False


posting_text_indicates_visa_sponsor_offer = posting_text_indicates_h1b_sponsor_offer


def _load_employer_index(cache_root: Path) -> dict[str, dict[str, Any]]:
    global _EMPLOYER_INDEX
    if _EMPLOYER_INDEX is not None:
        return _EMPLOYER_INDEX
    path = employer_index_path(cache_root)
    if not path.is_file():
        _EMPLOYER_INDEX = {}
        return _EMPLOYER_INDEX
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        _EMPLOYER_INDEX = {}
        return _EMPLOYER_INDEX
    employers = payload.get("employers") if isinstance(payload, dict) else None
    if isinstance(employers, dict):
        _EMPLOYER_INDEX = {str(k): v for k, v in employers.items() if isinstance(v, dict)}
    else:
        _EMPLOYER_INDEX = {}
    return _EMPLOYER_INDEX


def _pick_employer_record(company_name: str, index: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    target = normalize_employer_name(company_name)
    if not target:
        return None
    if target in index:
        return index[target]
    if len(target) < 4:
        return None
    candidates: list[tuple[int, dict[str, Any]]] = []
    for key, rec in index.items():
        if not key or len(key) < 4:
            continue
        if key.startswith(target) or target.startswith(key):
            candidates.append((abs(len(key) - len(target)), rec))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0])
    return candidates[0][1]


def lookup_company_h1b_meta(
    company_id: str,
    company_name: str,
    *,
    cache_root: Path,
) -> dict[str, Any]:
    cache_key = f"{company_id}:{company_name}"
    if cache_key in _COMPANY_LOOKUP_CACHE:
        return _COMPANY_LOOKUP_CACHE[cache_key]

    cached_path = company_h1b_cache_path(cache_root, company_id)
    if cached_path.is_file():
        try:
            data = json.loads(cached_path.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                _COMPANY_LOOKUP_CACHE[cache_key] = data
                return data
        except (OSError, json.JSONDecodeError):
            pass

    index = _load_employer_index(cache_root)
    rec = _pick_employer_record(company_name, index)
    if not rec:
        out = {
            "company_id": company_id,
            "company_name": company_name,
            "matched_employer": "",
            "lca_certified": 0,
            "lca_denied": 0,
            "grade": "F",
            "confidence": "low",
            "index_loaded": bool(index),
            "filer": False,
            "label": "No DOL visa filings",
        }
    else:
        certified = int(rec.get("lca_certified") or 0)
        denied = int(rec.get("lca_denied") or 0)
        grade, confidence = compute_h1b_grade(certified, denied)
        if rec.get("grade"):
            grade = str(rec["grade"])
        if rec.get("confidence"):
            confidence = str(rec["confidence"])
        matched = str(rec.get("display_name") or rec.get("employer_name") or "").strip()
        out = {
            "company_id": company_id,
            "company_name": company_name,
            "matched_employer": matched,
            "lca_certified": certified,
            "lca_denied": denied,
            "grade": grade,
            "confidence": confidence,
            "index_loaded": True,
            "filer": certified > 0,
            "label": f"Visa filer · {grade}" if certified > 0 else "No DOL visa filings",
        }

    try:
        cached_path.write_text(json.dumps(out, indent=2) + "\n", encoding="utf-8")
    except OSError:
        pass
    _COMPANY_LOOKUP_CACHE[cache_key] = out
    return out


def h1b_job_skip_reason(
    title: str,
    description_text: str,
    company_meta: dict[str, Any] | None = None,
    *,
    index_required: bool = True,
) -> str | None:
    """Return skip reason only when the posting explicitly denies visa sponsorship."""
    _ = company_meta, index_required
    blob = f"{title}\n{description_text}"
    if posting_text_indicates_no_visa_sponsorship(blob):
        return "posting explicitly denies visa sponsorship"
    return None


def company_is_dol_visa_filer(company_meta: dict[str, Any]) -> bool:
    return bool(company_meta.get("filer"))


def job_has_visa_sponsor_jd_signal(title: str, description_text: str) -> bool:
    """Positive visa sponsorship wording in the JD (board legend filter signal)."""
    blob = f"{title}\n{description_text}"
    return posting_text_indicates_h1b_sponsor_offer(blob)


def job_has_visa_sponsor_board_signal(
    title: str,
    description_text: str,
    company_meta: dict[str, Any] | None = None,
) -> bool:
    """Positive JD wording or known DOL filer (company badge / ancillary signal, not scrape gate)."""
    if job_has_visa_sponsor_jd_signal(title, description_text):
        return True
    return company_is_dol_visa_filer(company_meta or {})


def h1b_company_badge_html(grade: str, label: str) -> str:
    grade_key = str(grade or "F").strip().upper()[:1] or "F"
    text = str(label or "Visa").strip() or "Visa"
    return f'<span class="badge badge-h1b badge-h1b-{grade_key.lower()}">{text}</span>'


def _find_header_index(headers: list[str], *candidates: str) -> int | None:
    for candidate in candidates:
        key = candidate.upper()
        if key in headers:
            return headers.index(key)
    return None


def aggregate_lca_xlsx(
    path: Path,
    aggregates: dict[str, dict[str, Any]],
    wage_buckets: dict[str, dict[str, Any]] | None = None,
) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return 0
        headers = [str(cell or "").strip().upper() for cell in header_row]
        emp_i = _find_header_index(headers, "EMPLOYER_NAME", "EMPLOYER NAME")
        status_i = _find_header_index(headers, "CASE_STATUS", "CASE STATUS")
        visa_i = _find_header_index(headers, "VISA_CLASS", "VISA CLASS")
        title_i = _find_header_index(headers, "JOB_TITLE", "JOB TITLE")
        soc_i = _find_header_index(headers, "SOC_CODE", "SOC CODE", "SOC_CODE_ID")
        wage_from_i = _find_header_index(
            headers, "WAGE_RATE_OF_PAY_FROM", "WAGE RATE OF PAY FROM", "WAGE_FROM"
        )
        wage_to_i = _find_header_index(
            headers, "WAGE_RATE_OF_PAY_TO", "WAGE RATE OF PAY TO", "WAGE_TO"
        )
        wage_unit_i = _find_header_index(
            headers, "WAGE_UNIT_OF_PAY", "WAGE UNIT OF PAY", "WAGE_UNIT"
        )
        if emp_i is None or status_i is None:
            return 0
        seen = 0
        for row in rows:
            if not row or emp_i >= len(row):
                continue
            employer = str(row[emp_i] or "").strip()
            if not employer:
                continue
            if visa_i is not None and visa_i < len(row):
                visa = str(row[visa_i] or "").strip().upper()
                if visa and "H-1B" not in visa and visa not in {"H1B", "H-1B1"}:
                    continue
            status = str(row[status_i] or "").strip().upper()
            key = normalize_employer_name(employer)
            if not key:
                continue
            bucket = aggregates.setdefault(
                key,
                {
                    "display_name": employer,
                    "employer_name": employer,
                    "lca_certified": 0,
                    "lca_denied": 0,
                },
            )
            if status == "CERTIFIED":
                bucket["lca_certified"] += 1
            elif status in {"DENIED", "DENIED CERTIFICATION"}:
                bucket["lca_denied"] += 1
            seen += 1

            if wage_buckets is None or status != "CERTIFIED":
                continue
            if soc_i is not None and soc_i < len(row):
                soc = str(row[soc_i] or "").strip()
                # Prefer computer/math occupations (SOC 15-xxxx); keep unknown SOC.
                if soc and not (soc.startswith("15-") or soc.startswith("15")):
                    continue
            wage_from = row[wage_from_i] if wage_from_i is not None and wage_from_i < len(row) else None
            wage_to = row[wage_to_i] if wage_to_i is not None and wage_to_i < len(row) else None
            wage_unit = (
                row[wage_unit_i] if wage_unit_i is not None and wage_unit_i < len(row) else "Year"
            )
            annual = lca_row_annual_wage(wage_from, wage_to, wage_unit)
            if annual is None or annual < 40_000 or annual > 1_500_000:
                continue
            job_title = ""
            if title_i is not None and title_i < len(row):
                job_title = str(row[title_i] or "").strip()
            wage_rec = wage_buckets.setdefault(
                key,
                {
                    "display_name": employer,
                    "wages": [],
                    "wage_counts": {},
                    "by_title": {},
                },
            )
            _reservoir_add(
                wage_rec["wages"],
                wage_rec["wage_counts"],
                "all",
                float(annual),
                _WAGE_RESERVOIR_N,
            )
            for needle in _title_needles_for(job_title):
                title_bucket = wage_rec["by_title"].setdefault(
                    needle, {"wages": [], "wage_counts": {}}
                )
                _reservoir_add(
                    title_bucket["wages"],
                    title_bucket["wage_counts"],
                    needle,
                    float(annual),
                    _TITLE_RESERVOIR_N,
                )
        return seen
    finally:
        wb.close()


def download_dol_lca_xlsx(fy: int, quarter: int, dest: Path) -> None:
    url = DOL_LCA_URL.format(fy=fy, quarter=quarter)
    dest.parent.mkdir(parents=True, exist_ok=True)
    try:
        from curl_cffi import requests as curl_requests

        for impersonate in ("chrome120", "chrome124", "chrome"):
            try:
                resp = curl_requests.get(url, impersonate=impersonate, timeout=300)
            except Exception:
                continue
            if resp.status_code == 200 and resp.content:
                dest.write_bytes(resp.content)
                return
    except ImportError:
        pass
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept": "*/*",
        },
    )
    with urllib.request.urlopen(req, timeout=300) as resp:
        dest.write_bytes(resp.read())


def _finalize_wage_index(
    wage_buckets: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    employers: dict[str, dict[str, Any]] = {}
    for key, rec in wage_buckets.items():
        overall = _wage_summary(
            list(rec.get("wages") or []),
            sample_count=int((rec.get("wage_counts") or {}).get("all") or 0) or None,
        )
        by_title: dict[str, Any] = {}
        for needle, title_rec in (rec.get("by_title") or {}).items():
            if not isinstance(title_rec, dict):
                continue
            summary = _wage_summary(
                list(title_rec.get("wages") or []),
                sample_count=int((title_rec.get("wage_counts") or {}).get(needle) or 0)
                or None,
            )
            if summary and int(summary.get("n") or 0) >= _MIN_TITLE_SAMPLES:
                by_title[needle] = summary
        if not overall and not by_title:
            continue
        if overall and int(overall.get("n") or 0) < _MIN_EMPLOYER_SAMPLES and not by_title:
            continue
        employers[key] = {
            "display_name": rec.get("display_name") or key,
            "overall": overall,
            "by_title": by_title,
        }
    return employers


def _load_wage_index(cache_root: Path) -> dict[str, dict[str, Any]]:
    global _WAGE_INDEX
    path = wage_index_path(cache_root)
    if _WAGE_INDEX is not None:
        # Empty cache may mean the index was missing at first load; retry if present now.
        if _WAGE_INDEX or not path.is_file():
            return _WAGE_INDEX
        _WAGE_INDEX = None
    if not path.is_file():
        _WAGE_INDEX = {}
        return _WAGE_INDEX
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        _WAGE_INDEX = {}
        return _WAGE_INDEX
    employers = payload.get("employers") if isinstance(payload, dict) else None
    if isinstance(employers, dict):
        _WAGE_INDEX = {str(k): v for k, v in employers.items() if isinstance(v, dict)}
    else:
        _WAGE_INDEX = {}
    return _WAGE_INDEX


def wage_index_ready(cache_root: Path) -> bool:
    path = wage_index_path(cache_root)
    if not path.is_file():
        return False
    return bool(_load_wage_index(cache_root))


def lookup_lca_salary_range(
    company_name: str,
    job_title: str,
    *,
    cache_root: Path,
) -> dict[str, Any] | None:
    """Return DOL LCA wage range for employer (+ title needle when possible).

    Used when a posting does not disclose salary. Values are attested LCA wages
    (often for H-1B roles), not a guarantee of the open req's band.
    """
    index = _load_wage_index(cache_root)
    if not index:
        return None
    rec = _pick_employer_record(company_name, index)
    if not rec:
        return None
    by_title = rec.get("by_title") if isinstance(rec.get("by_title"), dict) else {}
    needles = _title_needles_for(job_title)
    summary: dict[str, Any] | None = None
    matched_title = ""
    for needle in needles:
        candidate = by_title.get(needle)
        if isinstance(candidate, dict) and int(candidate.get("n") or 0) >= _MIN_TITLE_SAMPLES:
            summary = candidate
            matched_title = needle
            break
    if summary is None:
        overall = rec.get("overall")
        if isinstance(overall, dict) and int(overall.get("n") or 0) >= _MIN_EMPLOYER_SAMPLES:
            summary = overall
            matched_title = ""
    if not summary:
        return None
    p25 = int(summary.get("p25") or 0)
    p75 = int(summary.get("p75") or 0)
    if p25 <= 0 or p75 <= 0:
        return None
    if p25 > p75:
        p25, p75 = p75, p25
    return {
        "matched_employer": str(rec.get("display_name") or company_name),
        "matched_title": matched_title,
        "n": int(summary.get("n") or 0),
        "p25": p25,
        "p50": int(summary.get("p50") or p25),
        "p75": p75,
        "label": format_lca_salary_label(p25, p75),
        "source": "dol_lca",
    }


def build_employer_index(
    cache_root: Path,
    *,
    quarters: list[tuple[int, int]] | None = None,
    progress: Callable[[str], None] | None = None,
) -> Path:
    quarters = quarters or [(2024, 4), (2025, 4)]
    aggregates: dict[str, dict[str, Any]] = {}
    wage_buckets: dict[str, dict[str, Any]] = {}
    raw_dir = h1b_index_dir(cache_root) / "dol"
    raw_dir.mkdir(parents=True, exist_ok=True)
    for fy, quarter in quarters:
        dest = raw_dir / f"LCA_Disclosure_Data_FY{fy}_Q{quarter}.xlsx"
        if progress:
            progress(f"Downloading FY{fy} Q{quarter}…")
        if not dest.is_file():
            download_dol_lca_xlsx(fy, quarter, dest)
        if progress:
            progress(f"Parsing {dest.name}…")
        aggregate_lca_xlsx(dest, aggregates, wage_buckets)

    employers: dict[str, dict[str, Any]] = {}
    for key, rec in aggregates.items():
        certified = int(rec.get("lca_certified") or 0)
        denied = int(rec.get("lca_denied") or 0)
        grade, confidence = compute_h1b_grade(certified, denied)
        employers[key] = {
            "display_name": rec.get("display_name") or rec.get("employer_name") or key,
            "lca_certified": certified,
            "lca_denied": denied,
            "grade": grade,
            "confidence": confidence,
        }

    out_path = employer_index_path(cache_root)
    payload = {
        "source": "dol_lca_disclosure",
        "quarters": [{"fy": fy, "quarter": q} for fy, q in quarters],
        "employer_count": len(employers),
        "employers": employers,
    }
    out_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    global _EMPLOYER_INDEX
    _EMPLOYER_INDEX = employers

    wage_employers = _finalize_wage_index(wage_buckets)
    wage_path = wage_index_path(cache_root)
    wage_payload = {
        "source": "dol_lca_disclosure_wages",
        "quarters": [{"fy": fy, "quarter": q} for fy, q in quarters],
        "employer_count": len(wage_employers),
        "notes": (
            "Certified H-1B/H-1B1 LCA wages (SOC 15-xxxx when present), annualized. "
            "p25–p75 ranges for salary badges when a posting omits pay. "
            "Not a guarantee of the open req's band."
        ),
        "employers": wage_employers,
    }
    wage_path.write_text(json.dumps(wage_payload, indent=2) + "\n", encoding="utf-8")
    global _WAGE_INDEX
    _WAGE_INDEX = wage_employers
    if progress:
        progress(f"Wrote wage index ({len(wage_employers)} employers) → {wage_path}")
    return out_path
